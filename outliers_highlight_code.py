# -*- coding: utf-8 -*-
"""
Created on Tue Jul 14 21:30:16 2020

@author: User
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

index_df = pd.read_excel(r'C:\Users\nimmi\PycharmProjects\untitled\Baidu_Scraping_Code\Merged_File\Baidu_2.5_output.xlsx')
keylist = list(set(index_df['Keyword'].tolist()))

finallist = []


def outlier_treatment(datacolumn, company):
    outlierlist = []
    sorted(datacolumn)
    Q1, Q3 = np.percentile(datacolumn, [25, 75])
    IQR = Q3 - Q1
    lower_range = Q1 - (1.5 * IQR)
    upper_range = Q3 + (1.5 * IQR)
    for ind in temp_df.index:
        if temp_df['Index'][ind] > upper_range or temp_df['Index'][ind] < lower_range:
            outlierlist.append(1)
        else:
            outlierlist.append(0)
    temp_df['outlier'] = outlierlist
    return temp_df


main_df = pd.DataFrame()

for k in keylist:
    temp_df = index_df[(index_df['Keyword'] == k)]
    main_df = main_df.append(outlier_treatment(temp_df['Index'], k))

main_df.to_excel(r'C:\Users\nimmi\PycharmProjects\untitled\Baidu_Scraping_Code\Merged_File\baidu_index_color.xlsx',
                 index=False)
wb = load_workbook(r'C:\Users\nimmi\PycharmProjects\untitled\Baidu_Scraping_Code\Merged_File\baidu_index_color.xlsx')

ws = wb['Sheet1']

redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')

start_col = 5
end_col = 5
for row_cells in ws.iter_rows():
    for cell in row_cells:
        if cell.value == 1 or cell.value == '1':
            for col in ['A', 'B', 'C', 'D']:
                color_cell = ws[str(col) + str(cell.coordinate).replace('D', '')]
                color_cell.fill = redFill

wb.save(r'C:\Users\nimmi\PycharmProjects\untitled\Baidu_Scraping_Code\Merged_File\baidu_index_highlighted.xlsx')